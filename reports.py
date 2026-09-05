'''
This module contains the report generation functions for the application.
'''

from os import path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from HospitalDatabase import PatientTable, DoctorTable
from datetime import datetime   
import CustomExceptions

def auto_size_columns(worksheet) ->None:
    """
    Auto-size all columns based on content
    """
    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                print(f"Error with cell {cell}: {e}")
                continue

        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width
        

def prepare_report_workbook(report_name:str, sheet_name:str, header_labels:list) -> Workbook:
    """
    Checks if a report file exists. If it does not exist, creates a new report file with the specified sheet name and headers.
    
    Args:
        report_name (str): The name of the report file  to check.
        sheet_name (str): The name of the sheet to check.
        header_labels (list): A list of strings representing the labels for each header.
        
    Returns:
        Workbook: The workbook object for the report.
    """

    #Check if the report file exists
    if path.exists(report_name):
        report_workbook = load_workbook(report_name)
        #Check if the sheet exists in the report file
        if sheet_name not in report_workbook.sheetnames:
            #Create a new sheet in the report file
            report_sheet = report_workbook.create_sheet(title=sheet_name)

            HEADER_COUNT = len(header_labels)
            #Set the headers for the new sheet
            for i in range(1, HEADER_COUNT + 1):
                report_sheet.cell(row=1, column=i).font = Font(bold=True)
                report_sheet.cell(row=1, column=i).alignment = Alignment(horizontal='center')

            #Append the header labels to the new sheet
            report_sheet.append(header_labels)
            
            report_workbook.save(report_name)
    
    else:
        #Create a new report file
        report_workbook = Workbook()
        report_sheet = report_workbook.create_sheet(title=sheet_name)

        HEADER_COUNT = len(header_labels)
        #Set the headers for the new sheet
        for i in range(1, HEADER_COUNT + 1):
            report_sheet.cell(row=1, column=i).font = Font(bold=True)
            report_sheet.cell(row=1, column=i).alignment = Alignment(horizontal='center')

        #Append the header labels to the new sheet
        report_sheet.append(header_labels)

        report_workbook.save(report_name)

    return report_workbook

def generate_patient_report() -> None:
    """
    Generates a patient report in Excel format.
    The report includes patient details such as ID, Name, Age, Disease and Assigned Doctor.
    """

    #Get the patient data from the database
    patient_data = PatientTable().view_patient_details()

    #Prepare the report workbook and sheet
    patient_workbook = prepare_report_workbook("patient_report.xlsx", "Patient Details", ["Patient ID", "Assigned Doctor", "Name", "Age", "Disease"])
    patient_sheet = patient_workbook["Patient Details"]
    last_row = patient_sheet.max_row
    if last_row <= 1:
        last_patient_id = None
    else:
        last_value = patient_sheet.cell(last_row, column=1).value
        try:
            last_patient_id = int(last_value) if last_value is not None else None
        except (ValueError, TypeError):
            last_patient_id = None

    #Write the patient data to the report sheet
    for patient in patient_data:
        if last_patient_id is None or patient[0] > last_patient_id:
            last_row += 1
            #Get the doctor name
            doctor_id = patient[1]
            doctor_data = DoctorTable().get_doctor_data_by_doctorid(doctor_id)
            doctor_name = doctor_data[1]
            
            patient_sheet.cell(row=last_row, column=1, value=patient[0])  # Patient ID
            patient_sheet.cell(row=last_row, column=2, value=doctor_name) # Assigned Doctor's Name
            patient_sheet.cell(row=last_row, column=3, value=patient[2])  # Name
            patient_sheet.cell(row=last_row, column=4, value=patient[3])  # Age
            patient_sheet.cell(row=last_row, column=5, value=patient[4])  # Disease

    auto_size_columns(patient_sheet)

    patient_workbook.save("patient_report.xlsx")


def generate_doctor_report() -> None:
    """
    Generates a doctor report in Excel format.
    The report includes doctor details such as ID, Name, Age, Specialization and Number of Patients.
    """

    #Get the doctor data from the database
    doctor_data = DoctorTable().view_doctor_details()

    #Prepare the report workbook and sheet
    doctor_workbook = prepare_report_workbook("doctor_report.xlsx", "Doctor Details", ["Doctor ID", "Name", "Specialization", "Created At"])
    doctor_sheet = doctor_workbook["Doctor Details"]
    last_row = doctor_sheet.max_row
    if last_row <= 1:
        last_doctor_id = None
    else:
        last_value = doctor_sheet.cell(last_row, column=1).value
        try:
            last_doctor_id = int(last_value) if last_value is not None else None
        except (ValueError, TypeError):
            last_doctor_id = None

    #Write the doctor data to the report sheet
    for doctor in doctor_data:
        if last_doctor_id is None or doctor[0] > last_doctor_id:
            last_row += 1
            doctor_sheet.cell(row=last_row, column=1, value=doctor[0])  # Doctor ID
            doctor_sheet.cell(row=last_row, column=2, value=doctor[1])  # Name
            doctor_sheet.cell(row=last_row, column=3, value=doctor[2])  # Specialisation
            doctor_sheet.cell(row=last_row, column=4, value=doctor[3])  # Created at

    auto_size_columns(doctor_sheet)
    doctor_workbook.save("doctor_report.xlsx")

def generate_summary_report() -> None:
    """
    Generates a summary report with multiple sheets:
    1. Overview - Total patients, total doctors, report date
    2. Patients per Doctor - Doctor name and patient count
    3. Disease Summary - Disease and count
    """

    # ==================== Sheet 1: Overview =============================
    #Prepare the Overview sheet
    overview_workbook = prepare_report_workbook('summary_workbook.xlsx', 'Overview', ['Total Patients', 'Total Doctors', 'Report Generated Date'] )
    overview_sheet = overview_workbook['Overview']

    NO_OF_PATIENTS, NO_OF_DOCTORS = len(PatientTable().view_patient_details()), len(DoctorTable().view_doctor_details())
    CREATED_AT = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    overview_sheet.append([NO_OF_PATIENTS,NO_OF_DOCTORS, CREATED_AT])

    auto_size_columns(overview_sheet)
    overview_workbook.save('summary_workbook.xlsx')

    # ==================== Sheet 2: Patients per Doctor ====================
    #Prepare the Patients per Doctor sheet
    patient_per_doctor_workbook = prepare_report_workbook('summary_workbook.xlsx','Patients per Doctor', ['Doctor Name', 'Assigned Patient Count'])
    patient_per_doctor_sheet = patient_per_doctor_workbook['Patients per Doctor']

    doctor_data = DoctorTable().view_doctor_details()

    DOCTOR_COUNT = len(doctor_data)
    #Getting patient per doctor count
    
    for i in range(1, DOCTOR_COUNT+1):
        data = doctor_data[i-1]
        doctor_id = data[0]
        doctor_name = data[1]
        try:
            patient_count = PatientTable().patient_per_doctor_count(doctor_id)
            patient_per_doctor_sheet.cell(row=i+1, column=2, value=patient_count[0])
                
        except CustomExceptions.RecordNotFoundError:
            patient_per_doctor_sheet.cell(row=i+1, column=2, value = 0)
            
        patient_per_doctor_sheet.cell(row=i+1, column=1, value=doctor_name)
        
        

    auto_size_columns(patient_per_doctor_sheet)
    patient_per_doctor_workbook.save('summary_workbook.xlsx')

    # ==================== Sheet 3: Disease Summary ========================
    #Prepare the Disease summary Report
    disease_summary_workbook = prepare_report_workbook('summary_workbook.xlsx', 'Disease Summary', ['Disease Name', 'Disease Count'])
    disease_summary_sheet = disease_summary_workbook['Disease Summary']

    patient_data = PatientTable().disease_count()

    PATIENT_COUNT = len(patient_data)
    for i in range(1, PATIENT_COUNT+1):
        data = patient_data[i-1]
        disease_summary_sheet.cell(row=i+1, column=1, value=data[0])  #Disease Name
        disease_summary_sheet.cell(row=i+1, column=2, value=data[1])  #Disease Count

    auto_size_columns(disease_summary_sheet)
    disease_summary_workbook.save('summary_workbook.xlsx')